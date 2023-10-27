import pandas as pd
import numpy as np
import math
import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

import tkinter as tk
from tkinter import filedialog
import easygui

easygui.msgbox("Choose your file to be proccessed (Excel File)")
#print("Asking for input file")
#root = tk.Tk()
#root.withdraw()

#file_path = filedialog.askopenfilename()

# Reemplaza 'nombre_del_archivo.xlsx' con el nombre de tu archivo Excel
excel_file = pd.ExcelFile('Script_input.xlsx')
easygui.msgbox("The file was processed successfully")

hojas ='Sheet1' #str(easygui.enterbox(msg="In which sheet do you want to work? "))
# Reemplaza 'nombre_de_la_hoja' con el nombre de la hoja en tu archivo Excel
df = excel_file.parse(hojas)


nombrecolumna1 = "Resource Name" #str(easygui.enterbox(msg="Put the name of the first column you want to work on (coworkers,name of the jobs) "))
nombrecolumna2 = "Sep"#str(easygui.enterbox(msg="Put the name of the second column you want to work on (hours) "))

# Reemplaza los valores nulos en la columna 2 con ceros
df[nombrecolumna2] = df[nombrecolumna2].fillna(0)
columna2 = df[nombrecolumna2]
columna1 = df[nombrecolumna1]
#easygui.msgbox("The names of the sheet and columns are correct")
df['EGB_Group'] = '' 


Nombres_Tareas = []
hours = []
Productivity = []
EGB_Group = []

cont = 0
conditional = 0
Productivity_abs = 0
Productivity_rel = 0
Productivity_sum = 0
Var_productivity = 0
cont_rel = 0

WORK_DAYS = 21#int(easygui.enterbox(msg="How many work days are in this report?: "))
#easygui.msgbox(WORK_DAYS, title="Días de trabajo ")

for index, row in df.iterrows():
    columna_tareas = row[nombrecolumna1]
    columna_horas = row[nombrecolumna2]
    

    if 'EGB' in columna_tareas:
        cont = cont + 1
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        conditional = 1

        if cont > 1:
            Productivity.append(Var_productivity/148)

            if Var_productivity != 0:
                cont_rel = cont_rel + 1

            Var_productivity = 0

            
        if 'EGB3' in columna_tareas:
            EGB_Group.append('EGB3')
        elif 'EGB8' in columna_tareas:
            EGB_Group.append('EGB8')
        elif 'EGB9' in columna_tareas:
            EGB_Group.append('EGB9')
        elif 'EGB10' in columna_tareas:
            EGB_Group.append('EGB10')
        elif 'EGB11' in columna_tareas:
            EGB_Group.append('EGB11')
        elif 'EGB12' in columna_tareas:
            EGB_Group.append('EGB12')
        else:
            EGB_Group.append(' ')

    elif columna_tareas.startswith("   P_") and conditional == 1 and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')
        Productivity_sum = Productivity_sum + columna_horas
        Var_productivity = Var_productivity + columna_horas

    
    elif conditional == 1  and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')

    
    elif not 'EGB' in columna_tareas and (columna_tareas.endswith('-MS)') or columna_tareas.endswith('-MX)') or columna_tareas.endswith('-SX)')):
        conditional = 0
    
Productivity.append(Var_productivity/148)




###########################################################
#                   Archivo Output
# #########################################################
    
  
#easygui.msgbox("Put the name for you output file")
#archivos = str(easygui.enterbox(msg="Put a name to the output file "))  
#archivo  = archivos + ".xlsx"
archivo = "Salida.xlsx"

directorio = "C:\\Users\OJA5GA\Documents\VS_Code\V1"
#easygui.msgbox("enter the folder where you want the output file to be stored ")
#directorio=str(easygui.diropenbox())
mes = nombrecolumna2
HOURS_PER_DAY = 8.25

ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)

# Verificar si el archivo existe

if not os.path.exists(ruta_completa):
    workbook = openpyxl.Workbook()      #Crear un nuevo libro
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

    easygui.msgbox(f"The file {archivo} was created in: {directorio}")
    #easygui.msgbox("This file was created")
else:
    #easygui.msgbox("This file already exists")
    easygui.msgbox(f"The file {archivo} already exists in: {directorio}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

#Imprimir encabezados
hoja['A1'] = 'Resource Name'
hoja['B1'] = 'Hours'
hoja['C1'] = 'Productivity'
hoja['D1'] = 'Comments'
hoja['E1'] = 'EGB Group'
hoja['G1'] = 'Days'
hoja['G2'] = WORK_DAYS
hoja['H1'] = 'Hr x day'
hoja['H2'] = HOURS_PER_DAY
hoja['I1'] = 'Total hrs'
TOTAL_HOURS = HOURS_PER_DAY*WORK_DAYS
hoja['I2'] = TOTAL_HOURS

Productivity_abs = (Productivity_sum/148) / cont
Productivity_rel = (Productivity_sum/148) / cont_rel


#Resultados
#hoja['G5'] = 'Abs. Productivity Sum(%)'
#hoja['G6'] = Productivity_sum/148
hoja['G5'] = 'Qty People'
hoja['G6'] = cont
hoja['H5'] = 'Average Abs. Productivity'
hoja['H6'] = Productivity_abs
hoja['I5'] = 'Average Rel. Productivity'
hoja['I6'] = Productivity_rel
hoja['H9'] = 'High Value (+10%)'
HIGH_VALUE = TOTAL_HOURS*1.1
hoja['H10'] = HIGH_VALUE
hoja['I9'] = 'Low Value (-5%)'
LOW_VALUE = TOTAL_HOURS*0.95
hoja['I10'] = LOW_VALUE



#Imprimir nombres y tareas
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    hoja.cell(row=filas+1, column=1, value=contenido)

#Imprimir horas
for filas, contenido in enumerate(hours, start=1):
    hoja.cell(row=filas+1, column=2, value=contenido)

#Imprimir EGB group
for filas, contenido in enumerate(EGB_Group, start=1):
    hoja.cell(row=filas+1, column=5, value=contenido)


#Imprimir Productividad
counter = 0
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    if 'EGB' in contenido:
       hoja.cell(row=filas+1, column=3, value=Productivity[counter])
       counter +=1  
    


####### ESTABLECER ESTILO DE PORCENTAJE EN ALGUNAS CELDAS ##########
# Verificar si el estilo de porcentaje ya existe
porcentaje_style = None
for style in workbook.style_names:
    if style == 'porcentaje_style':
        porcentaje_style = style
        break

# Si el estilo no existe, se crea
if porcentaje_style is None:
    porcentaje_style = openpyxl.styles.NamedStyle(name='porcentaje_style')
    porcentaje_style.number_format = '0.00%'

columna = 3
filas_1 = 2

# Aplicar el estilo a la columna de productividad
for fila in hoja.iter_rows(min_row=2, min_col=columna, max_col=columna, values_only=True):
    celda = hoja.cell(row=filas_1, column=columna)
    celda.style = porcentaje_style
    filas_1 += 1

celda = hoja['H6'] 
celda.style = porcentaje_style

celda = hoja['I6'] 
celda.style = porcentaje_style


####################################################
################# COLOR DE CELDAS ##################
####################################################

# Define los colores de relleno
relleno_verde__titulos = PatternFill(start_color="41ad2c", end_color="41ad2c", fill_type="solid")
relleno_verde = PatternFill(start_color="7ce12e", end_color="7ce12e", fill_type="solid")
relleno_amarillo = PatternFill(start_color="d9e221", end_color="d9e221", fill_type="solid")
relleno_rojo = PatternFill(start_color="f25757", end_color="f25757", fill_type="solid")
relleno_gris = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

relleno_egb = PatternFill(start_color="ff9e9e", end_color="ff9e9e", fill_type="solid")
relleno_egb3 = PatternFill(start_color="ffd69e", end_color="ffd69e", fill_type="solid")
relleno_egb8 = PatternFill(start_color="faffa1", end_color="faffa1", fill_type="solid")
relleno_egb9 = PatternFill(start_color="c9ff92", end_color="c9ff92", fill_type="solid")
relleno_egb10 = PatternFill(start_color="92e6ff", end_color="92e6ff", fill_type="solid")
relleno_egb11 = PatternFill(start_color="c77bff", end_color="c77bff", fill_type="solid")
relleno_egb12 = PatternFill(start_color="ff7bd6", end_color="ff7bd6", fill_type="solid")

#Defenir el estilo de letras
negritas = Font(bold = True)
letras_rojas = Font(color="FF0000")

#### Aplicar los estilos a los Títulos ########
#Encabezados
for i in range (9):
    celda = hoja.cell(row = 1, column = i+1)
    celda.fill = relleno_verde__titulos
    celda.font = negritas

#Productivity Headers
for i in range (3):
    celda = hoja.cell(row = 5, column = i+7)
    celda.fill = relleno_verde__titulos
    celda.font = negritas

#High and low values headers
for i in range (2):
    celda = hoja.cell(row = 9, column = i+8)
    celda.fill = relleno_verde__titulos
    celda.font = negritas

#Days, hours per days, total hrs
for i in range (3):
    celda = hoja.cell(row = 2, column = i+7)
    celda.fill = relleno_amarillo

#Productivity Results
for i in range (3):
    celda = hoja.cell(row = 6, column = i+7)
    celda.fill = relleno_amarillo
    celda.font = negritas

#High Value result
celda = hoja['H10']
celda.fill = relleno_verde
celda.font = negritas

#Low value result
celda = hoja['I10']
celda.fill = relleno_rojo
celda.font = negritas


###### Cambiar el formato de nombres de asociados, horas, productividad y grupo ######
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    celda_nombre = hoja.cell(row=filas+1, column=1)
    celda_horas = hoja.cell(row=filas+1, column=2)
    celda_prod = hoja.cell(row=filas+1, column=3)
    celda_egb = hoja.cell(row=filas+1, column=5)


    if "EGB" in contenido:
        #Aplica formato a los nombres de asociado
        celda_nombre.fill = relleno_gris        
        celda_nombre.font = negritas
        celda_nombre.font = letras_rojas

        #Aplica formato a la suma de horas
        if celda_horas.value >= HIGH_VALUE:
            celda_horas.fill = relleno_verde
        elif celda_horas.value < LOW_VALUE:
            celda_horas.fill = relleno_rojo
        else:
            celda_horas.fill = relleno_amarillo
        celda_horas.font = negritas

        #Aplica formato a las productividades
        if celda_prod.value >= 0.85:
            celda_prod.fill = relleno_verde
        elif celda_prod.value < 0.75:
            celda_prod.fill = relleno_rojo
        else:
            celda_prod.fill = relleno_amarillo
        celda_prod.font = negritas


        #Aplica formato a la columna de grupos
        if 'EGB3' in celda_egb.value:
            celda_egb.fill = relleno_egb3
            celda_egb.font = negritas
            EGB_Group.append('EGB3')
        elif 'EGB8' in celda_egb.value:
            celda_egb.fill = relleno_egb8
            celda_egb.font = negritas
        elif 'EGB9' in celda_egb.value:
            celda_egb.fill = relleno_egb9
            celda_egb.font = negritas
        elif 'EGB10' in celda_egb.value:
            celda_egb.fill = relleno_egb10
            celda_egb.font = negritas
        elif 'EGB11' in celda_egb.value:
            celda_egb.fill = relleno_egb11
            celda_egb.font = negritas
        elif 'EGB12' in celda_egb.value:
            celda_egb.fill = relleno_egb12
            celda_egb.font = negritas
        else:
            celda_egb.fill = relleno_egb
            celda_egb.font = negritas
       

# CREAR TABLA
existing_tables = hoja.tables
table_name = 'Tabla_general'
 
if table_name in existing_tables:
    # Si la tabla ya existe, la eliminamos
    hoja._tables.remove(existing_tables[table_name])
 
 
#Crea una nueva tabla
table = Table(displayName=table_name, ref="A:E")  
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=False, showColumnStripes=False)
table.tableStyleInfo = style
hoja.add_table(table)



# Ajustar el ancho de las columnas de la hoja de excel
for columna in hoja.columns:
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda

    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 2

workbook.save(ruta_completa)

if os.path.exists(ruta_completa):
    easygui.msgbox("Save completed, opening file...")
    os.startfile(ruta_completa)
    #os.system(f'start excel "{ruta_completa}"')
else:
    easygui.msgbox(f'El archivo "{ruta_completa}" no existe.')









